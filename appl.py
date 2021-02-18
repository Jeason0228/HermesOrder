# -*- coding: utf-8 -*-
import asyncio
import json
from datetime import datetime
from tkinter.messagebox import *
from tkinter import scrolledtext, Tk, ttk, filedialog
import settings
import aiohttp
import jsonpath
from mttkinter import mtTkinter as mtk
import threading
import time
import hmac
import hashlib
import base64
import urllib.parse
import requests
from openpyxl import load_workbook
import sys
import multiprocessing
import os

asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
import warnings

warnings.filterwarnings("ignore")


class Application:

    def __init__(self, master):
        self.root = master
        self.root.geometry("500x480")
        self.root.title("Hermes")
        self.__createUI()
        self.status = False
        self.addLog("初始化成功...")

    def __createUI(self):
        '''
        创建界面
        :return:
        '''
        # 数据查询
        self.searchBox = mtk.LabelFrame(self.root, text="商品查询", fg="blue")
        self.searchBox.place(x=20, y=20, width=460, height=130)

        self.searchLink = mtk.Label(self.searchBox, text="商品链接：")
        self.searchLink.place(x=20, y=10, width=70, height=25)
        self.searchLinkText = mtk.Text(self.searchBox)
        self.searchLinkText.place(x=90, y=10, width=350, height=75)
        self.searchLinkText.bind(sequence="<Double-Button-1>", func=lambda x: self.thread_it(self.exportExcel))

        # 日志
        self.logs = mtk.LabelFrame(self.root, text="log日志", fg="blue")
        self.logs.place(x=20, y=170, width=200, height=295)
        self.logtext = scrolledtext.ScrolledText(self.logs, fg="green")
        self.logtext.place(x=10, y=5, width=180, height=255)

        self.settings = mtk.LabelFrame(self.root, text="设置", fg="blue")
        self.settings.place(x=240, y=170, width=240, height=245)

        self.proxy = mtk.Label(self.settings, text="选择代理：")
        self.proxy.place(x=10, y=15, width=90, height=25)
        self.proxyText = ttk.Combobox(self.settings)
        self.proxyText["values"] = list(settings.PROXIES.keys())
        self.proxyText.place(x=100, y=15, width=110, height=25)

        self.user = mtk.Label(self.settings, text="选择账号：")
        self.user.place(x=10, y=65, width=90, height=25)
        self.userText = ttk.Combobox(self.settings)
        self.userText["values"] = list(settings.HERMES_INFO.keys())
        self.userText.place(x=100, y=65, width=110, height=25)

        self.dingding = mtk.Label(self.settings, text="钉钉群：")
        self.dingding.place(x=10, y=105, width=90, height=25)
        self.dingdingText = ttk.Combobox(self.settings)
        self.dingdingText["values"] = list(settings.DINGDING.keys())
        self.dingdingText.place(x=100, y=105, width=110, height=25)

        self.netLan = mtk.Label(self.settings, text="国家代码：")
        self.netLan.place(x=10, y=145, width=90, height=25)
        self.netText = mtk.Entry(self.settings)
        self.netText.place(x=100, y=145, width=110, height=25)

        self.email = mtk.Label(self.settings, text="消息标题：")
        self.email.place(x=10, y=185, width=90, height=25)
        self.emailText = mtk.Entry(self.settings)
        self.emailText.place(x=100, y=185, width=110, height=25)

        # self.timeSleep = mtk.Label(self.settings, text="延时（秒）：")
        # self.timeSleep.place(x=10, y=170, width=90, height=25)
        # self.timeSleepText = mtk.Entry(self.settings)
        # self.timeSleepText.place(x=100, y=170, width=110, height=25)

        self.startBtn = mtk.Button(self.root, text="启动", command=lambda: self.thread_it(self.start))
        self.startBtn.place(x=240, y=430, width=100, height=35)

        self.stopBtn = mtk.Button(self.root, text="退出", command=lambda: self.thread_it(self.stop))
        self.stopBtn.place(x=370, y=430, width=100, height=35)

    def sendMessage(self, title, sendText, dingding):
        try:
            secret = settings.DINGDING[dingding]["secret"]
            webHook = settings.DINGDING[dingding]["webHook"]
            # 使用钉钉机器人发送定制消息
            timestamp = str(round(time.time() * 1000))
            secret_enc = secret.encode('utf-8')
            string_to_sign = '{}\n{}'.format(timestamp, secret)
            string_to_sign_enc = string_to_sign.encode('utf-8')
            hmac_code = hmac.new(secret_enc, string_to_sign_enc, digestmod=hashlib.sha256).digest()
            sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))
            webhook = "{}&timestamp={}&sign={}".format(webHook, timestamp, sign)
            headers = {"Content-Type": "application/json",
                       "Charset": "UTF-8"}
            # 消息类型和数据格式参照钉钉开发文档
            data = {
                "actionCard": {
                    "title": title,
                    "text": title,
                    "btnOrientation": "0",
                    "singleTitle": "查看商品详情" if sendText else "",
                    "singleURL": sendText,
                },
                "msgtype": "actionCard"
            }

            r = requests.post(webhook, data=json.dumps(data), headers=headers)
            time.sleep(1)
        except Exception as e:
            print(e)

    async def __getContent(self, semaphore, link_, proxy, emailTitle, link, dingding):

        # 代理服务器
        proxyHost = "http-dyn.abuyun.com"
        proxyPort = "9020"

        # 代理隧道验证信息
        proxyUser = settings.PROXIES[proxy]["proxyUser"]
        proxyPass = settings.PROXIES[proxy]["proxyPass"]

        proxyServer = "http://%(user)s:%(pass)s@%(host)s:%(port)s" % {
            "host": proxyHost,
            "port": proxyPort,
            "user": proxyUser,
            "pass": proxyPass,
        }
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Referer': 'https://www.hermes.com/',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Safari/537.36',
            "origin": "https://www.hermes.com",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-site"
        }
        conn = aiohttp.TCPConnector(verify_ssl=False)
        async with semaphore:
            async with aiohttp.ClientSession(headers=headers, connector=conn, trust_env=True) as session:
                try:
                    async with session.get(link_, proxy=proxyServer, timeout=4) as resp:

                        if str(resp.status).startswith("2"):
                            content = await resp.json()
                            skuList = jsonpath.jsonpath(content, "$..variants..sku")
                            sku = content.get("sku")
                            if sku in skuList:
                                self.sendMessage(emailTitle, link, dingding)
                                self.status = True
                                return True, link
                        return False, False
                except Exception as e:
                    print(e)
                    return False, False

    async def taskManager(self, linkList, link_list, emailTitle, proxy, dingding):
        tasks = []
        semaphore = asyncio.Semaphore(settings.SEMNUM)
        for index, link_ in enumerate(link_list):
            link = linkList[index]
            task = asyncio.ensure_future(self.__getContent(semaphore, link_, proxy, emailTitle, link, dingding))
            task.add_done_callback(self.callback)
            tasks.append(task)
        await asyncio.gather(*tasks)

    def callback(self, feature):
        result, link = feature.result()
        if not result:
            self.addLog("无货,继续刷新!")
        else:
            self.addLog("有货!有货!")
            self.order.make_order(link)

    @staticmethod
    def thread_it(func, *args):
        t = threading.Thread(target=func, args=args)
        t.setDaemon(True)
        t.start()

    def addLog(self, msg):
        self.logtext.insert(mtk.END, "{} {}\n".format(datetime.now().strftime("%H:%M:%S"), msg))
        self.logtext.yview_moveto(1.0)

    def exportExcel(self):
        excelPath = filedialog.askopenfilename(title=u'选择文件')
        if not excelPath.endswith(".xlsx"):
            showerror("错误信息", "Excel格式错误.")
        self.excelData = []
        wb = load_workbook(excelPath)
        ws = wb.active
        products = [str(i[0]) for i in list(ws.values)[1:] if i]
        totals = ws.max_row - 1
        showinfo("提示信息", f"共获取到{totals}条商品链接.")
        self.searchLinkText.insert("end", "\n".join(products).strip())

    def start(self):
        from orderM import OrderProject
        self.order = OrderProject()
        self.addLog("浏览器初始化成功...")
        linkList = self.searchLinkText.get('0.0', 'end').split("\n")
        if not linkList:
            return showerror("错误信息", "未获取到商品链接.")

        skuList = [link_.split('-')[-1].strip("/").replace("p", "-") for link_ in linkList]

        proxy = self.proxyText.get().strip()
        if not proxy:
            showerror("警告信息", "请设置代理.")
            return
        # 账号信息
        userInfo = self.userText.get().strip()
        if not userInfo:
            showerror("错误信息", "请选择购买账号.")
            return
        username = settings.HERMES_INFO.get(userInfo).get("user")
        passwd = settings.HERMES_INFO.get(userInfo).get("passwd")
        self.order.login(username, passwd)

        dingding = self.dingdingText.get().strip()
        if not dingding:
            showerror("警告信息", "请选择钉钉群.")
            return

        emailTitle = self.emailText.get().strip()
        if not emailTitle:
            showerror("错误信息", "请输入消息标题!")
            return

        net = self.netText.get().strip()
        if not net or "/" not in net:
            showerror("错误信息", "请输入正确的国家代码!")
            return

        link_list = [
            f"https://bck.hermes.cn/product?locale={net.replace('/', '_')}&productsku={sku.replace('v', '%2520')}" for
            sku in skuList if sku]
        timeNow = datetime.now()
        self.addLog("开始监控商品状态..")
        while True:
            new_loop = asyncio.new_event_loop()
            asyncio.set_event_loop(new_loop)
            self.loop = asyncio.get_event_loop()
            self.loop.run_until_complete(self.taskManager(linkList, link_list, emailTitle, proxy, dingding))
            total_time = (datetime.now() - timeNow).total_seconds() / 60
            if total_time > settings.RESTARTSLEEPTIME:
                self.order.restartBrowser()
                timeNow = datetime.now()

            if self.status:
                time.sleep(settings.SLEEPTIME)

    def stop(self):
        self.addLog("当前任务已停止!!")
        os._exit(0)


if __name__ == '__main__':
    if sys.platform.startswith('win'):
        multiprocessing.freeze_support()
    root = Tk()
    Application(root)
    root.mainloop()
